import io
import os
from datetime import datetime, date
import subprocess
from apps.workflow.models import ActionRequests
from pathlib import Path
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4
from reportlab.lib.utils import ImageReader
from django.views.decorators.clickjacking import xframe_options_sameorigin
from django.conf import settings
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.utils import timezone
from django.db import transaction

from django.utils.text import get_valid_filename
from django.http import HttpResponse, FileResponse, Http404
from django.db.models import Q
from django.core.paginator import Paginator
from django.utils.dateparse import parse_datetime, parse_date
from django.http import JsonResponse
from django.template.loader import render_to_string
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

from apps.records.models import Records, RecordAttachments,ExternalCompanyNames
from apps.organization.models import Departments
from apps.workflow.models import AuditLogs


# ============================================================
# Internal aliases (MUST be defined early; helpers use these)
# ============================================================
Record = Records
RecordAttachment = RecordAttachments
Department = Departments
AuditLog = AuditLogs


# ============================================================
# Auth helpers (session-based) + role checks (Admin/Clerk)
# ============================================================

def _parse_dt_local(s: str):
    """
    Parses <input type="datetime-local"> value: 'YYYY-MM-DDTHH:MM'
    Returns a datetime or None.
    """
    s = (s or "").strip()
    if not s:
        return None
    dt = parse_datetime(s)
    if not dt:
        return None
    # If your project uses timezone-aware datetimes, make it aware
    if getattr(settings, "USE_TZ", False) and timezone.is_naive(dt):
        dt = timezone.make_aware(dt, timezone.get_current_timezone())
    return dt

def _get_or_create_external_company_name(name: str, description: str = ""):
    name = (name or "").strip()
    description = (description or "").strip()

    if not name:
        return None

    existing = ExternalCompanyNames.objects.filter(CompanyName__iexact=name).first()
    if existing:
        return existing

    now = timezone.now()
    return ExternalCompanyNames.objects.create(
        CompanyName=name,
        Description=description or None,
        CreatedAt=now,
        UpdatedAt=now,
    )
    
    
def _parse_d(s: str):
    """
    Parses <input type="date"> value: 'YYYY-MM-DD'
    Returns a date or None.
    """
    s = (s or "").strip()
    return parse_date(s) if s else None

def _get_current_user(request):
    user_id = request.session.get("user_id")
    if not user_id:
        return None
    from apps.accounts.models import Users
    try:
        return Users.objects.get(UserID=user_id, IsActive=True)
    except Users.DoesNotExist:
        return None


def _get_or_create_department_by_name(name: str, description: str = ""):
    """
    Create Department by name if it doesn't exist; otherwise return existing.
    Only new incoming departments should get description copied from the record.
    """
    name = (name or "").strip()
    description = (description or "").strip()

    if not name:
        return None

    dept = Department.objects.filter(DepartmentName__iexact=name).first()
    if dept:
        return dept

    return Department.objects.create(
        DepartmentName=name,
        Description=description or None,
        CreatedAt=timezone.now()
    )
    

def _user_in_group(user, group_name: str) -> bool:
    if not user:
        return False
    from apps.accounts.models import UserGroups
    return UserGroups.objects.filter(
        UserID=user,
        GroupID__GroupName__iexact=group_name
    ).exists()
    

def _record_form_context(request, user, mode="create", record=None, attachments=None):
    departments = Department.objects.all().order_by("DepartmentName")

    form_data = {
        "MessengerName": (request.POST.get("MessengerName") or "").strip(),
        "InvoiceNumber": (request.POST.get("InvoiceNumber") or "").strip(),
        "Subject": (request.POST.get("Subject") or "").strip(),
        "DateReceived": (request.POST.get("DateReceived") or "").strip(),
        "ExternalDocument": (request.POST.get("ExternalDocument") or "No").strip(),
        "external_company_name": (request.POST.get("external_company_name") or "").strip(),
        "Description": (request.POST.get("Description") or "").strip(),
        "IncomingDepartmentID": (request.POST.get("IncomingDepartmentID") or "").strip(),
        "incoming_department_new": (request.POST.get("incoming_department_new") or "").strip(),
        "OutgoingDepartmentID": (request.POST.get("OutgoingDepartmentID") or "").strip(),
        "outgoing_department_new": (request.POST.get("outgoing_department_new") or "").strip(),
        "DateDispatched": (request.POST.get("DateDispatched") or "").strip(),
        "Status": (request.POST.get("Status") or "").strip(),
        "Returned": (request.POST.get("Returned") or "").strip(),
        "DateReturned": (request.POST.get("DateReturned") or "").strip(),
    }

    return {
        "mode": mode,
        "record": record,
        "attachments": attachments or [],
        "departments": departments,
        "is_admin": _is_admin(user),
        "is_clerk": _is_clerk(user),
        "form_data": form_data or {},
    }
    

def _build_filtered_records_qs(request, user):
    q = (request.GET.get("q") or "").strip()
    status = (request.GET.get("Status") or "").strip()
    invoice_number_raw = (request.GET.get("invoice_number") or "").strip()
    subject = (request.GET.get("subject") or "").strip()
    messenger_name = (request.GET.get("messenger_name") or "").strip()
    returned = (request.GET.get("returned") or "").strip()

    # CORRECTION: read pending request filter
    pending_request = (request.GET.get("pending_request") or "").strip()

    incoming_department_id = (request.GET.get("incoming_department_id") or "").strip()
    external_company_name_id = (request.GET.get("external_company_name_id") or "").strip()
    outgoing_department_id = (request.GET.get("outgoing_department_id") or "").strip()
    include_deleted = request.GET.get("include_deleted") == "1" and _is_admin(user)

    qs = Record.objects.all()

    # =========================
    # Date ranges
    # =========================
    date_received_from = _parse_dt_local(request.GET.get("date_received_from"))
    date_received_to = _parse_dt_local(request.GET.get("date_received_to"))
    if date_received_from:
        qs = qs.filter(DateReceived__gte=date_received_from)
    if date_received_to:
        qs = qs.filter(DateReceived__lte=date_received_to)

    date_dispatched_from = _parse_dt_local(request.GET.get("date_dispatched_from"))
    date_dispatched_to = _parse_dt_local(request.GET.get("date_dispatched_to"))
    if date_dispatched_from:
        qs = qs.filter(DateDispatched__gte=date_dispatched_from)
    if date_dispatched_to:
        qs = qs.filter(DateDispatched__lte=date_dispatched_to)

    date_returned_from = _parse_d(request.GET.get("date_returned_from"))
    date_returned_to = _parse_d(request.GET.get("date_returned_to"))
    if date_returned_from:
        qs = qs.filter(DateReturned__gte=date_returned_from)
    if date_returned_to:
        qs = qs.filter(DateReturned__lte=date_returned_to)

    # =========================
    # Simple filters
    # =========================
    if returned in ("Yes", "No") and hasattr(Record, "Returned"):
        qs = qs.filter(Returned=returned)

    # CORRECTION: apply pending workflow request filter
    if pending_request:
        pending_qs = ActionRequests.objects.filter(Status__iexact="pending")

        if pending_request == "any":
            pending_ids = pending_qs.values_list("TargetRecordID", flat=True)
            qs = qs.filter(RecordID__in=pending_ids)

        elif pending_request == "none":
            pending_ids = pending_qs.values_list("TargetRecordID", flat=True)
            qs = qs.exclude(RecordID__in=pending_ids)

        elif pending_request in ("EDIT", "DELETE", "RESTORE"):
            pending_ids = pending_qs.filter(
                RequestType__iexact=pending_request
            ).values_list("TargetRecordID", flat=True)
            qs = qs.filter(RecordID__in=pending_ids)

    # Invoice exact
    if invoice_number_raw and hasattr(Record, "InvoiceNumber"):
        invoice_digits = invoice_number_raw.replace(",", "").strip()
        if invoice_digits.isdigit():
            qs = qs.filter(InvoiceNumber__startswith=invoice_digits)

    if subject and hasattr(Record, "Subject"):
        qs = qs.filter(Subject__icontains=subject)

    if messenger_name and hasattr(Record, "MessengerName"):
        qs = qs.filter(MessengerName__icontains=messenger_name)

    if status and hasattr(Record, "Status"):
        qs = qs.filter(Status=status)

    # =========================
    # Soft delete visibility
    # =========================
    if not include_deleted:
        qs = qs.filter(**_record_is_deleted_filter())

    # =========================
    # Department filters (FK)
    # =========================
    if hasattr(Record, "ExternalCompanyName"):
        if external_company_name_id == "__has_value__":
            qs = qs.filter(ExternalCompanyName__isnull=False)
        elif external_company_name_id.isdigit():
            qs = qs.filter(ExternalCompanyName_id=int(external_company_name_id))
        
    if incoming_department_id.isdigit() and hasattr(Record, "IncomingDepartmentID"):
        qs = qs.filter(IncomingDepartmentID_id=int(incoming_department_id))

    if outgoing_department_id.isdigit() and hasattr(Record, "OutgoingDepartmentID"):
        qs = qs.filter(OutgoingDepartmentID_id=int(outgoing_department_id))

    # =========================
    # Search (q)
    # =========================
    if q:
        filters = Q()

        q_digits = q.replace(",", "").strip()
        if q_digits.isdigit() and hasattr(Record, "InvoiceNumber"):
            filters |= Q(InvoiceNumber=int(q_digits))

        for field in ["MessengerName", "Subject", "Description"]:
            if hasattr(Record, field):
                filters |= Q(**{f"{field}__icontains": q})
            
        if hasattr(Record, "ExternalCompanyName"):
                filters |= Q(ExternalCompanyName__CompanyName__icontains=q)

        qs = qs.filter(filters)

    # =========================
    # Sorting
    # =========================
    sort = (request.GET.get("sort") or "").strip().lower()
    dir_ = (request.GET.get("dir") or "desc").strip().lower()

    SORT_MAP = {
        "received": "DateReceived",
        "dispatched": "DateDispatched",
        "returned": "DateReturned",
        "invoice": "InvoiceNumber",
        "messenger": "MessengerName",
        "subject": "Subject",
    }

    sort_field = SORT_MAP.get(sort)
    if sort_field:
        prefix = "" if dir_ == "asc" else "-"
        qs = qs.order_by(f"{prefix}{sort_field}", "-RecordID")
    else:
        qs = qs.order_by("-RecordID")

    return qs

def _has_any_filter_params(request) -> bool:
    FILTER_KEYS = [
        # text/flags
        "q", "Status", "invoice_number", "subject", "messenger_name", "returned",
        "external_company_name_id",

        # CORRECTION: include pending_request as its own filter key
        "pending_request",

        "incoming_department_id", "outgoing_department_id", "include_deleted",

        # date ranges
        "date_received_from", "date_received_to",
        "date_dispatched_from", "date_dispatched_to",
        "date_returned_from", "date_returned_to",

        # sorting
        "sort", "dir",
    ]
    for k in FILTER_KEYS:
        v = (request.GET.get(k) or "").strip()
        if v:
            return True
    return False

def _is_admin(user) -> bool:
    return _user_in_group(user, "Admin")


def _is_clerk(user) -> bool:
    return _user_in_group(user, "Clerk")


def _can_edit_records(user) -> bool:
    return _is_admin(user) or _is_clerk(user)


def login_required_view(view_func):
    def wrapper(request, *args, **kwargs):
        user = _get_current_user(request)
        if not user:
            return redirect("accounts:login")
        request.current_user = user
        return view_func(request, *args, **kwargs)
    return wrapper


def clerk_or_admin_required_view(view_func):
    def wrapper(request, *args, **kwargs):
        user = _get_current_user(request)
        if not user:
            return redirect("accounts:login")
        if not _can_edit_records(user):
            return redirect("accounts:access_denied")
        request.current_user = user
        return view_func(request, *args, **kwargs)
    return wrapper


def _audit(request, action: str, details: str = "",target_record=None):
    user = _get_current_user(request)
    
    if user is None:
        return
    
    # try:
    data = {
        "ActorUserID":user,
        "EventType":action,
        "Details":details,
        "EventTime":timezone.now(),
    }
    if target_record is not None:
        data["TargetRecordID"] = target_record
    
    AuditLog.objects.create(**data)
    #except Exception:
        #pass

def _get_pending_delete_request_for_record(record):
    return ActionRequests.objects.filter(
        TargetRecordID=record,
        RequestType__iexact="DELETE",
        Status__iexact="pending",
    ).order_by("-RequestID").first()

# ============================================================
# Utilities for schema alignment (PascalCase fields)
# ============================================================
def _record_is_deleted_filter():
    """
    Return a filter that selects NON-deleted records.
    """
    if hasattr(Record, "IsDeleted"):
        return {"IsDeleted": False}
    if hasattr(Record, "is_deleted"):
        return {"is_deleted": False}
    if hasattr(Record, "deleted_at"):
        return {"deleted_at__isnull": True}
    if hasattr(Record, "is_active"):
        return {"is_active": True}
    return {}


def _mark_record_deleted(rec, user=None):
    """
    Soft delete: prefer IsDeleted=True if it exists.
    """
    now = timezone.now()
    if hasattr(rec, "IsDeleted"):
        rec.IsDeleted = True
    if hasattr(rec, "is_deleted"):
        rec.is_deleted = True
    if hasattr(rec, "deleted_at"):
        rec.deleted_at = now
    # ✅ schema-safe: your Users PK is UserID, not id
    if hasattr(rec, "deleted_by_id") and user:
        rec.deleted_by_id = getattr(user, "UserID", None)
    if hasattr(rec, "is_active"):
        rec.is_active = False
    rec.save()


def _attachment_is_deleted_filter():
    if hasattr(RecordAttachment, "IsDeleted"):
        return {"IsDeleted": False}
    if hasattr(RecordAttachment, "is_deleted"):
        return {"is_deleted": False}
    if hasattr(RecordAttachment, "deleted_at"):
        return {"deleted_at__isnull": True}
    return {}


def _mark_attachment_deleted(att, user=None):
    now = timezone.now()
    if hasattr(att, "IsDeleted"):
        att.IsDeleted = True
    if hasattr(att, "is_deleted"):
        att.is_deleted = True
    if hasattr(att, "deleted_at"):
        att.deleted_at = now
    # ✅ schema-safe: your Users PK is UserID, not id
    if hasattr(att, "deleted_by_id") and user:
        att.deleted_by_id = getattr(user, "UserID", None)
    att.save()


def _parse_datetime_local(value: str):
    """
    Parse HTML <input type="datetime-local"> value: 'YYYY-MM-DDTHH:MM' (seconds optional).
    Returns timezone-aware datetime if possible.
    """
    value = (value or "").strip()
    if not value:
        return None
    try:
        dt = datetime.fromisoformat(value)
        if timezone.is_naive(dt):
            try:
                dt = timezone.make_aware(dt, timezone.get_current_timezone())
            except Exception:
                pass
        return dt
    except Exception:
        return None


def _parse_date(value: str):
    value = (value or "").strip()
    if not value:
        return None
    try:
        return date.fromisoformat(value)
    except Exception:
        return None


# ============================================================
# File storage helpers for attachments (schema-aligned)
# ============================================================
def _ensure_media_root():
    return bool(getattr(settings, "MEDIA_ROOT", None))


def _save_uploaded_file(record_id: int, uploaded_file):
    """
    Save file under MEDIA_ROOT/attachments/<RecordID>/filename
    Returns relative path to store in FilePath.
    """
    safe_name = get_valid_filename(uploaded_file.name)
    rel_dir = os.path.join("attachments", str(record_id))
    abs_dir = os.path.join(settings.MEDIA_ROOT, rel_dir)
    os.makedirs(abs_dir, exist_ok=True)

    rel_path = os.path.join(rel_dir, safe_name)
    abs_path = os.path.join(settings.MEDIA_ROOT, rel_path)

    # avoid overwrite collisions
    base, ext = os.path.splitext(safe_name)
    counter = 1
    while os.path.exists(abs_path):
        safe_name2 = f"{base}_{counter}{ext}"
        rel_path = os.path.join(rel_dir, safe_name2)
        abs_path = os.path.join(settings.MEDIA_ROOT, rel_path)
        counter += 1

    with open(abs_path, "wb") as out:
        for chunk in uploaded_file.chunks():
            out.write(chunk)

    return rel_path


# ============================================================
# 1) home_view
# ============================================================
@login_required_view
def home_view(request):
    user = request.current_user
    return render(request, "records/home.html", {
        "is_admin": _is_admin(user),
        "is_clerk": _is_clerk(user),
    })


# ============================================================
# 2) dashboard_view
# ============================================================
@login_required_view
def dashboard_view(request):
    user = request.current_user

    # ✅ FIX: Total = all records (deleted + non-deleted)
    total_records = Record.objects.all().count()

    # ✅ FIX: Active = non-deleted records (soft delete aware)
    active_records = Record.objects.filter(**_record_is_deleted_filter()).count()

    # ✅ FIX: Deleted = soft-deleted records
    deleted_records = Record.objects.filter(IsDeleted=True).count() if hasattr(Record, "IsDeleted") else 0

    total_departments = Department.objects.count()
    total_attachments = RecordAttachment.objects.filter(**_attachment_is_deleted_filter()).count()

    # ✅ FIX: Recent records should show active (non-deleted) records
    recent_records = Record.objects.filter(**_record_is_deleted_filter()).order_by("-RecordID")[:10]
    recent_audit = AuditLog.objects.all().order_by("-EventTime")[:10]

    return render(request, "records/dashboard.html", {
        "total_records": total_records,
        "total_departments": total_departments,
        "total_attachments": total_attachments,
        "recent_records": recent_records,
        "recent_audit": recent_audit,
        "is_admin": _is_admin(user),
        "is_clerk": _is_clerk(user),
        "active_records": active_records,
        "deleted_records":deleted_records,
    })


# ============================================================
# 3) records_table_view — (kept, but made schema-safe)
# ============================================================
@login_required_view
def records_table_view(request):
    user = request.current_user

    # =========================
    # CORRECTION: read values for template state only
    # Actual filtering now comes from _build_filtered_records_qs(...)
    # so pending_request works properly.
    # =========================
    q = (request.GET.get("q") or "").strip()
    status = (request.GET.get("Status") or "").strip()
    sort = (request.GET.get("sort") or "").strip().lower()
    dir_ = (request.GET.get("dir") or "desc").strip().lower()
    external_company_name_id = (request.GET.get("external_company_name_id") or "").strip()
    external_company_names = ExternalCompanyNames.objects.all().order_by("CompanyName")

    invoice_number_raw = (request.GET.get("invoice_number") or "").strip()
    subject = (request.GET.get("subject") or "").strip()
    messenger_name = (request.GET.get("messenger_name") or "").strip()
    returned = (request.GET.get("returned") or "").strip()

    # CORRECTION: keep selected pending request filter in UI
    pending_request = (request.GET.get("pending_request") or "").strip()

    incoming_department_id = (request.GET.get("incoming_department_id") or "").strip()
    outgoing_department_id = (request.GET.get("outgoing_department_id") or "").strip()
    include_deleted = request.GET.get("include_deleted") == "1" and _is_admin(user)

    # =========================
    # CORRECTION: use shared helper so all filters
    # including pending_request are applied consistently
    # =========================
    qs = _build_filtered_records_qs(request, user)

    # =========================
    # Pagination
    # =========================
    page_number = request.GET.get("page") or 1
    paginator = Paginator(qs, 25)
    page_obj = paginator.get_page(page_number)

    # =========================
    # CORRECTION: build stable pending request map
    # one request badge per record, newest pending first
    # =========================
    pending_requests_map = {}
    pending_requests_qs = ActionRequests.objects.filter(
        Status__iexact="pending"
    ).order_by("-RequestID")

    for req in pending_requests_qs:
        if req.TargetRecordID_id and req.TargetRecordID_id not in pending_requests_map:
            pending_requests_map[req.TargetRecordID_id] = (req.RequestType or "").upper()

    departments = Department.objects.all().order_by("DepartmentName")

    # =========================
    # Keep filters in UI
    # =========================
    ctx_filters = {
        "date_received_from": request.GET.get("date_received_from") or "",
        "date_received_to": request.GET.get("date_received_to") or "",
        "date_dispatched_from": request.GET.get("date_dispatched_from") or "",
        "date_dispatched_to": request.GET.get("date_dispatched_to") or "",
        "date_returned_from": request.GET.get("date_returned_from") or "",
        "date_returned_to": request.GET.get("date_returned_to") or "",
    }

    return render(request, "records/records_table.html", {
        "q": q,
        "sort": sort,
        "dir": dir_,
        "pending_requests_map": pending_requests_map,
        "Status": status,
        "invoice_number": invoice_number_raw,
        "subject": subject,
        "messenger_name": messenger_name,
        "returned": returned,
        "external_company_name_id": external_company_name_id,
        "external_company_names": external_company_names,

        # CORRECTION: pass pending_request back to template
        "pending_request": pending_request,

        "incoming_department_id": incoming_department_id,
        "outgoing_department_id": outgoing_department_id,
        "include_deleted": include_deleted,
        "page_obj": page_obj,
        "records": page_obj,
        "departments": departments,
        "is_admin": _is_admin(user),
        "is_clerk": _is_clerk(user),

        # date inputs
        **ctx_filters,
})
# ============================================================
# 4) record_detail_view
# ============================================================
@login_required_view
def record_detail_view(request, record_id):
    user = request.current_user
    rec = get_object_or_404(Record, RecordID=record_id)

    # CORRECTION: allow admins and clerks to view deleted records
    if hasattr(rec, "IsDeleted") and rec.IsDeleted:
        if not (_is_admin(user) or _is_clerk(user)):
            return redirect("accounts:access_denied")

    # CORRECTION: detect whether this detail page was opened from deleted records
    from_deleted = request.GET.get("from") == "deleted"

    attachments = RecordAttachment.objects.filter(
        RecordID_id=rec.RecordID,
        **_attachment_is_deleted_filter()
    ).order_by("-UploadedAt")

    latest_attachment = attachments.first()
    history = AuditLog.objects.all().order_by("-EventTime")[:20]

    # CORRECTION: clerks must not edit deleted records
    can_edit = _can_edit_records(user)
    if _is_clerk(user) and hasattr(rec, "IsDeleted") and rec.IsDeleted:
        can_edit = False

    return render(request, "records/record_detail.html", {
        "record": rec,
        "attachments": attachments,
        "latest_attachment": latest_attachment,
        "history": history,
        "can_edit": can_edit,
        "is_admin": _is_admin(user),
        "is_clerk": _is_clerk(user),
        "from_deleted": from_deleted,  # CORRECTION
    })


# ============================================================
# 5) record_create_view (Clerk/Admin) — UPDATED FOR YOUR FIELDS + inline dept create
# ============================================================
@clerk_or_admin_required_view
def record_create_view(request):
    user = request.current_user

    if request.method == "GET":
        return render(
            request,
            "records/record_form.html",
            _record_form_context(request, user, mode="create")
        )

    # ============================================================
    # 1) Read form fields
    # ============================================================
    messenger_name = (request.POST.get("MessengerName") or "").strip()
    subject = (request.POST.get("Subject") or "").strip()
    description = (request.POST.get("Description") or "").strip()
    status = (request.POST.get("Status") or "").strip()

    external_document = (request.POST.get("ExternalDocument") or "No").strip()
    external_company_name_raw = (request.POST.get("external_company_name") or "").strip()

    invoice_number_raw = (request.POST.get("InvoiceNumber") or "").strip()
    invoice_number_clean = invoice_number_raw.replace(",", "").strip()

    date_received = _parse_datetime_local(request.POST.get("DateReceived"))
    date_dispatched = _parse_datetime_local(request.POST.get("DateDispatched"))
    returned = (request.POST.get("Returned") or "").strip()
    date_returned = _parse_date(request.POST.get("DateReturned"))

    incoming_id = (request.POST.get("IncomingDepartmentID") or "").strip()
    outgoing_id = (request.POST.get("OutgoingDepartmentID") or "").strip()
    incoming_new = (request.POST.get("incoming_department_new") or "").strip()
    outgoing_new = (request.POST.get("outgoing_department_new") or "").strip()

    # ============================================================
    # 2) Validate required fields FIRST
    #    Under no circumstance should anything be created before this passes
    # ============================================================
    if not messenger_name:
        messages.error(request, "Messenger name is required.")
        return render(
            request,
            "records/record_form.html",
            _record_form_context(request, user, mode="create")
        )

    if not subject:
        messages.error(request, "Subject is required.")
        return render(
            request,
            "records/record_form.html",
            _record_form_context(request, user, mode="create")
        )

    if not description:
        messages.error(request, "Description is required.")
        return render(
            request,
            "records/record_form.html",
            _record_form_context(request, user, mode="create")
        )

    if not invoice_number_clean:
        messages.error(request, "Invoice number is required.")
        return render(
            request,
            "records/record_form.html",
            _record_form_context(request, user, mode="create")
        )

    if not status:
        messages.error(request, "Status is required.")
        return render(
            request,
            "records/record_form.html",
            _record_form_context(request, user, mode="create")
        )

    if not date_received:
        messages.error(request, "Date received is required (use the date picker).")
        return render(
            request,
            "records/record_form.html",
            _record_form_context(request, user, mode="create")
        )

    try:
        invoice_number_int = int(invoice_number_clean)
    except Exception:
        messages.error(request, "Invoice number must be numeric (commas allowed).")
        return render(
            request,
            "records/record_form.html",
            _record_form_context(request, user, mode="create")
        )

    # ============================================================
    # 3) Resolve departments WITHOUT creating anything yet
    # ============================================================
    incoming_dept = None
    outgoing_dept = None

    if external_document == "Yes":
        # External document: departments are not required
        incoming_id = ""
        outgoing_id = ""
        incoming_new = ""
        outgoing_new = ""
    else:
        # Resolve existing departments only
        if incoming_id:
            incoming_dept = Department.objects.filter(DepartmentID=incoming_id).first()

        if outgoing_id:
            outgoing_dept = Department.objects.filter(DepartmentID=outgoing_id).first()

        # Do not create new departments yet — just validate the names for now
        if not incoming_dept and not incoming_new:
            messages.error(request, "Incoming Department is required unless this is an external document.")
            return render(
                request,
                "records/record_form.html",
                _record_form_context(request, user, mode="create")
            )

    # ============================================================
    # 4) Business rules — validate everything before creating anything
    # ============================================================
    status_norm = status.lower()
    is_with_md = (status_norm == "with md")

    if date_dispatched and not outgoing_dept and not outgoing_new and external_document != "Yes":
        messages.error(request, "Outgoing Department is required when Date Dispatched is filled.")
        return render(
            request,
            "records/record_form.html",
            _record_form_context(request, user, mode="create")
        )

    if date_dispatched and is_with_md:
        if returned != "Yes":
            messages.error(request, "Returned must be 'Yes' when Status is 'With MD' and Date Dispatched is filled.")
            return render(
                request,
                "records/record_form.html",
                _record_form_context(request, user, mode="create")
            )

        if not date_returned:
            messages.error(request, "Date Returned is required when Status is 'With MD' and Date Dispatched is filled.")
            return render(
                request,
                "records/record_form.html",
                _record_form_context(request, user, mode="create")
            )

    # Optional extra consistency rule
    if date_returned and returned != "Yes":
        messages.error(request, "Returned must be 'Yes' when Date Returned is filled.")
        return render(
            request,
            "records/record_form.html",
            _record_form_context(request, user, mode="create")
        )

    # ============================================================
    # 5) Only NOW create related objects + record, inside one transaction
    # ============================================================
    try:
        with transaction.atomic():
            # Create/resolve departments only after all validation passes
            if external_document != "Yes":
                if not incoming_dept and incoming_new:
                    incoming_dept = _get_or_create_department_by_name(
                        incoming_new,
                        description=description
                    )

                if not outgoing_dept and outgoing_new:
                    outgoing_dept = _get_or_create_department_by_name(
                        outgoing_new
                    )

            # Create external company only after validation succeeds
            external_company_name = _get_or_create_external_company_name(
                external_company_name_raw,
                description=description
            )

            create_kwargs = dict(
                MessengerName=messenger_name,
                Subject=subject,
                Description=description,
                DateReceived=date_received,
                InvoiceNumber=invoice_number_int,
                IncomingDepartmentID=incoming_dept,
                OutgoingDepartmentID=outgoing_dept,
                ExternalDocument=external_document if external_document in ("Yes", "No") else "No",
                ExternalCompanyName=external_company_name,
                DateDispatched=date_dispatched,
                Returned=returned if returned in ("Yes", "No") else None,
                DateReturned=date_returned,
                Status=status,
            )

            if hasattr(Record, "CreatedAt"):
                create_kwargs["CreatedAt"] = timezone.now()
            if hasattr(Record, "UpdatedAt"):
                create_kwargs["UpdatedAt"] = timezone.now()

            rec = Record.objects.create(**create_kwargs)

            files = request.FILES.getlist("files")
            if files and _ensure_media_root():
                for f in files:
                    rel_path = _save_uploaded_file(rec.RecordID, f)
                    RecordAttachment.objects.create(
                        RecordID_id=rec.RecordID,
                        FilePath=rel_path,
                        OriginalFileName=f.name,
                        UploadedAt=timezone.now(),
                        UploadedByUserID=user,
                    )

            _audit(request, "RECORD_CREATE", f"Created record RecordID={rec.RecordID}", target_record=rec)

    except Exception:
        messages.error(request, "Record could not be created. Please correct the form and try again.")
        return render(
            request,
            "records/record_form.html",
            _record_form_context(request, user, mode="create")
        )

    messages.success(request, "Record created successfully.")
    return redirect("records:record_edit", record_id=rec.RecordID)

# ============================================================
# 6) record_edit_view (Clerk/Admin) — UPDATED FOR YOUR FIELDS + inline dept create
# ============================================================
@clerk_or_admin_required_view
def record_edit_view(request, record_id):
    user = request.current_user
    rec = get_object_or_404(Record, RecordID=record_id)

    # Prevent non-admin from viewing/editing deleted record
    if hasattr(rec, "IsDeleted") and rec.IsDeleted and not _is_admin(user):
        return redirect("accounts:access_denied")

    departments = Department.objects.all().order_by("DepartmentName")

    # ============================================================
    # GET: show form + attachments
    # ============================================================
    if request.method == "GET":
        attachments = RecordAttachment.objects.filter(
            RecordID_id=rec.RecordID,
            **_attachment_is_deleted_filter()
        ).order_by("-AttachmentID")

        return render(request, "records/record_form.html", {
            "mode": "edit",
            "record": rec,
            "departments": departments,
            "attachments": attachments,
            "is_admin": _is_admin(user),
            "is_clerk": _is_clerk(user),
        })

    # ============================================================
    # POST: read all values first
    # ============================================================
    messenger_name = (request.POST.get("MessengerName") or "").strip()
    subject = (request.POST.get("Subject") or "").strip()
    description = (request.POST.get("Description") or "").strip()
    status = (request.POST.get("Status") or "").strip()

    invoice_number_raw = (request.POST.get("InvoiceNumber") or "").strip()
    invoice_number_clean = invoice_number_raw.replace(",", "").strip()

    external_document = (request.POST.get("ExternalDocument") or "No").strip()
    external_company_name_raw = (request.POST.get("external_company_name") or "").strip()

    date_received = _parse_dt_local(request.POST.get("DateReceived"))
    date_dispatched = _parse_dt_local(request.POST.get("DateDispatched"))
    returned = (request.POST.get("Returned") or "").strip()
    date_returned = _parse_d(request.POST.get("DateReturned"))

    incoming_id = (request.POST.get("IncomingDepartmentID") or "").strip()
    outgoing_id = (request.POST.get("OutgoingDepartmentID") or "").strip()
    incoming_new = (request.POST.get("incoming_department_new") or "").strip()
    outgoing_new = (request.POST.get("outgoing_department_new") or "").strip()

    ALLOWED_STATUS = {"With MD", "Not with MD"}
    ALLOWED_RETURNED = {"", "Yes", "No"}

    attachments = RecordAttachment.objects.filter(
        RecordID_id=rec.RecordID,
        **_attachment_is_deleted_filter()
    ).order_by("-AttachmentID")

    def render_edit_with_error(msg):
        messages.error(request, msg)

        # Create a lightweight object carrying submitted values
        class TempRecord:
            pass

        temp = TempRecord()
        temp.RecordID = rec.RecordID
        temp.MessengerName = messenger_name
        temp.Subject = subject
        temp.Description = description
        temp.InvoiceNumber = invoice_number_raw
        temp.DateReceived = date_received
        temp.DateDispatched = date_dispatched
        temp.Returned = returned
        temp.DateReturned = date_returned
        temp.Status = status
        temp.ExternalDocument = external_document

        # Preserve selected/typed external company for form re-render
        temp.ExternalCompanyName = getattr(rec, "ExternalCompanyName", None)

        # Preserve selected departments if possible
        incoming_existing = Department.objects.filter(DepartmentID=incoming_id).first() if incoming_id else None
        outgoing_existing = Department.objects.filter(DepartmentID=outgoing_id).first() if outgoing_id else None

        temp.IncomingDepartmentID = incoming_existing
        temp.OutgoingDepartmentID = outgoing_existing

        return render(request, "records/record_form.html", {
            "mode": "edit",
            "record": temp,
            "departments": departments,
            "attachments": attachments,
            "is_admin": _is_admin(user),
            "is_clerk": _is_clerk(user),
            "incoming_department_new_value": incoming_new,
            "outgoing_department_new_value": outgoing_new,
            "external_company_name_value": external_company_name_raw,
        })

    # ============================================================
    # 1. Core required fields
    # ============================================================
    if not messenger_name:
        return render_edit_with_error("Messenger name is required.")

    if not subject:
        return render_edit_with_error("Subject is required.")

    if not description:
        return render_edit_with_error("Description is required.")

    if not invoice_number_clean:
        return render_edit_with_error("Invoice number is required.")

    if not status:
        return render_edit_with_error("Status is required.")

    if not date_received:
        return render_edit_with_error("Date received is required.")

    # ============================================================
    # 2. Invoice number rules
    # ============================================================
    try:
        invoice_number_int = int(invoice_number_clean)
    except Exception:
        return render_edit_with_error("Invoice number must be numeric. Commas are allowed in input.")

    if invoice_number_int <= 0:
        return render_edit_with_error("Invoice number must be greater than zero.")

    duplicate_qs = Record.objects.filter(InvoiceNumber=invoice_number_int).exclude(RecordID=rec.RecordID)
    if hasattr(Record, "IsDeleted"):
        duplicate_qs = duplicate_qs.filter(IsDeleted=False)

    if duplicate_qs.exists():
        return render_edit_with_error("A record with this invoice number already exists.")

    # ============================================================
    # 3. Choice validation
    # ============================================================
    if status not in ALLOWED_STATUS:
        return render_edit_with_error("Status must be either 'With MD' or 'Not with MD'.")

    if returned not in ALLOWED_RETURNED:
        return render_edit_with_error("Returned must be blank, 'Yes', or 'No'.")

    if external_document not in ("Yes", "No"):
        external_document = "No"

    # ============================================================
    # 4. External document / company rules
    # ============================================================
    if external_document == "Yes":
        if not external_company_name_raw:
            return render_edit_with_error("External company name is required for external documents.")

        incoming_id = ""
        outgoing_id = ""
        incoming_new = ""
        outgoing_new = ""

    # ============================================================
    # 5. Resolve existing departments only (do NOT create yet)
    # ============================================================
    incoming_dept = None
    outgoing_dept = None

    if external_document != "Yes":
        if incoming_id:
            incoming_dept = Department.objects.filter(DepartmentID=incoming_id).first()

        if outgoing_id:
            outgoing_dept = Department.objects.filter(DepartmentID=outgoing_id).first()

        if not incoming_dept and not incoming_new:
            return render_edit_with_error("Incoming Department is required unless this is an external document.")

    # ============================================================
    # 6. Dispatch / outgoing department rules
    # ============================================================
    if date_dispatched:
        if external_document != "Yes" and not outgoing_dept and not outgoing_new:
            return render_edit_with_error("Outgoing Department is required when Date Dispatched is filled.")

        if returned == "":
            return render_edit_with_error("Returned must be selected when Date Dispatched is filled.")

    # ============================================================
    # 7. Returned / date returned rules
    # ============================================================
    if returned == "Yes" and not date_returned:
        return render_edit_with_error("Date Returned is required when Returned is 'Yes'.")

    if returned == "No" and date_returned:
        return render_edit_with_error("Date Returned must be empty when Returned is 'No'.")

    if date_returned and returned != "Yes":
        return render_edit_with_error("Returned must be 'Yes' when Date Returned is filled.")

    # ============================================================
    # 8. Status-specific rule
    # ============================================================
    if status == "With MD" and date_dispatched:
        if returned != "Yes":
            return render_edit_with_error("Returned must be 'Yes' when Status is 'With MD' and Date Dispatched is filled.")

        if not date_returned:
            return render_edit_with_error("Date Returned is required when Status is 'With MD' and Date Dispatched is filled.")

    # ============================================================
    # 9. Date order rules
    # ============================================================
    if date_dispatched and date_dispatched < date_received:
        return render_edit_with_error("Date Dispatched cannot be earlier than Date Received.")

    if date_returned:
        received_date = date_received.date() if hasattr(date_received, "date") else date_received

        if date_returned < received_date:
            return render_edit_with_error("Date Returned cannot be earlier than Date Received.")

        if date_dispatched:
            dispatched_date = date_dispatched.date() if hasattr(date_dispatched, "date") else date_dispatched
            if date_returned < dispatched_date:
                return render_edit_with_error("Date Returned cannot be earlier than Date Dispatched.")

    # ============================================================
    # 10. Capture OLD values before any change
    # ============================================================
    old_messenger_name = rec.MessengerName
    old_subject = rec.Subject
    old_description = rec.Description
    old_date_received = rec.DateReceived
    old_invoice_number = rec.InvoiceNumber
    old_external_document = getattr(rec, "ExternalDocument", "") or ""
    old_external_company_name = getattr(getattr(rec, "ExternalCompanyName", None), "CompanyName", "") or ""
    old_incoming_department = getattr(rec.IncomingDepartmentID, "DepartmentName", "") if rec.IncomingDepartmentID else ""
    old_outgoing_department = getattr(rec.OutgoingDepartmentID, "DepartmentName", "") if rec.OutgoingDepartmentID else ""
    old_date_dispatched = rec.DateDispatched
    old_returned = rec.Returned
    old_date_returned = rec.DateReturned
    old_status = rec.Status

    # ============================================================
    # 11. Prepare resolved names for request or save
    # ============================================================
    incoming_existing_name = getattr(incoming_dept, "DepartmentName", "") if incoming_dept else ""
    outgoing_existing_name = getattr(outgoing_dept, "DepartmentName", "") if outgoing_dept else ""

    final_incoming_name = incoming_existing_name or incoming_new
    final_outgoing_name = outgoing_existing_name or outgoing_new

    # ============================================================
    # 12. Clerk edits become approval requests only
    # ============================================================
    if _is_clerk(user):
        def fmt_dt(v):
            if not v:
                return ""
            try:
                return v.strftime("%Y-%m-%d %H:%M")
            except Exception:
                return str(v)

        def fmt_date(v):
            if not v:
                return ""
            try:
                return v.strftime("%Y-%m-%d")
            except Exception:
                return str(v)

        change_lines = [
            "EDIT REQUEST",
            f"RecordID={rec.RecordID}",
            "",
            f"CHANGE|MessengerName|{old_messenger_name or ''}|{messenger_name}",
            f"CHANGE|Subject|{old_subject or ''}|{subject}",
            f"CHANGE|Description|{old_description or ''}|{description}",
            f"CHANGE|DateReceived|{fmt_dt(old_date_received)}|{fmt_dt(date_received)}",
            f"CHANGE|InvoiceNumber|{old_invoice_number if old_invoice_number is not None else ''}|{invoice_number_int}",
            f"CHANGE|IncomingDepartment|{old_incoming_department}|{final_incoming_name}",
            f"CHANGE|OutgoingDepartment|{old_outgoing_department}|{final_outgoing_name}",
            f"CHANGE|DateDispatched|{fmt_dt(old_date_dispatched)}|{fmt_dt(date_dispatched)}",
            f"CHANGE|Returned|{old_returned or ''}|{returned if returned in ('Yes', 'No') else ''}",
            f"CHANGE|DateReturned|{fmt_date(old_date_returned)}|{fmt_date(date_returned)}",
            f"CHANGE|Status|{old_status or ''}|{status}",
            f"CHANGE|ExternalDocument|{old_external_document}|{external_document}",
            f"CHANGE|ExternalCompanyName|{old_external_company_name}|{external_company_name_raw}",
        ]

        req = ActionRequests.objects.create(
            RequestType="EDIT",
            TargetRecordID=rec,
            RequestedByUserID=user,
            Status="pending",
            CreatedAt=timezone.now(),
            RequestDetails="\n".join(change_lines),
        )

        _audit(
            request,
            "REQUEST_CREATE_EDIT",
            f"Clerk requested edit for RecordID={rec.RecordID} (RequestID={req.RequestID})",
            target_record=rec,
        )

        messages.success(request, "Edit request submitted for admin approval.")
        return redirect("workflow:requests_list")

    # ============================================================
    # 13. Admin edits apply immediately
    # ============================================================
    try:
        with transaction.atomic():
            # Create departments only after all validation passes
            if external_document != "Yes":
                if not incoming_dept and incoming_new:
                    incoming_dept = _get_or_create_department_by_name(
                        incoming_new,
                        description=description
                    )

                if not outgoing_dept and outgoing_new:
                    outgoing_dept = _get_or_create_department_by_name(
                        outgoing_new
                    )

            # Create external company only after all validation passes
            external_company_name = None
            if external_document == "Yes":
                external_company_name = _get_or_create_external_company_name(
                    external_company_name_raw,
                    description=description
                )

            rec.MessengerName = messenger_name
            rec.Subject = subject
            rec.Description = description
            rec.DateReceived = date_received
            rec.InvoiceNumber = invoice_number_int
            rec.IncomingDepartmentID = incoming_dept
            rec.OutgoingDepartmentID = outgoing_dept
            rec.ExternalDocument = external_document
            rec.ExternalCompanyName = external_company_name
            rec.DateDispatched = date_dispatched
            rec.Returned = returned if returned in ("Yes", "No") else None
            rec.DateReturned = date_returned
            rec.Status = status

            if hasattr(rec, "UpdatedAt"):
                rec.UpdatedAt = timezone.now()

            rec.save()

            files = request.FILES.getlist("files")
            if files and _ensure_media_root():
                for f in files:
                    rel_path = _save_uploaded_file(rec.RecordID, f)
                    RecordAttachment.objects.create(
                        RecordID_id=rec.RecordID,
                        FilePath=rel_path,
                        OriginalFileName=f.name,
                        UploadedAt=timezone.now(),
                        UploadedByUserID=user,
                    )

            _audit(
                request,
                "RECORD_EDIT",
                f"Admin edited record RecordID={rec.RecordID}",
                target_record=rec
            )

    except Exception:
        return render_edit_with_error("Record could not be updated. Please correct the form and try again.")

    messages.success(request, "Record updated successfully.")
    return redirect("records:record_edit", record_id=rec.RecordID)
# ============================================================
# 7) record_soft_delete_view (Clerk/Admin)
# ============================================================
@clerk_or_admin_required_view
def record_soft_delete_view(request, record_id):
    user = request.current_user
    rec = get_object_or_404(Record, RecordID=record_id)

    if hasattr(rec, "IsDeleted") and rec.IsDeleted:
        messages.info(request, "Record is already deleted.")
        return redirect("records:record_detail", record_id=rec.RecordID)


    if _is_clerk(user):
        
        existing_pending = _get_pending_delete_request_for_record(rec)
        if existing_pending:
                messages.info(request, "A pending delete request already exists for this record.")
                return redirect("workflow:requests_list")

        req = ActionRequests.objects.create(
            RequestType="DELETE",
            TargetRecordID=rec,
            RequestedByUserID=user,
            RequestDetails=f"Request to delete RecordID={rec.RecordID}",
            Status="pending",
            CreatedAt=timezone.now(),
        )

        _audit(
            request,
            "REQUEST_CREATE_DELETE",
            f"Clerk requested delete for RecordID={rec.RecordID} (RequestID={req.RequestID})",
            target_record=rec,
        )

        messages.success(request, "Delete request submitted for admin approval.")
        return redirect("workflow:requests_list")




    _mark_record_deleted(rec, user=user)

    _audit(
        request,
        "RECORD_SOFT_DELETE",
        f"Admin soft-deleted RecordID={rec.RecordID}",
        target_record=rec
    )

    messages.success(request, "Record moved to deleted records.")
    return redirect("records:records_table")


# ============================================================
# CORRECTION: records_bulk_delete_view (Clerk/Admin)
# Admin  -> immediate soft delete for selected records
# Clerk  -> create one DELETE request per selected record
# ============================================================
@clerk_or_admin_required_view
def records_bulk_delete_view(request):
    user = request.current_user

    if request.method != "POST":
        return redirect("records:records_table")

    # CORRECTION: read selected record IDs from bulk form
    raw_ids = request.POST.getlist("record_ids")
    record_ids = []

    for value in raw_ids:
        value = str(value or "").strip()
        if value.isdigit():
            record_ids.append(int(value))

    # CORRECTION: remove duplicates while keeping only valid numeric IDs
    record_ids = list(dict.fromkeys(record_ids))

    if not record_ids:
        messages.error(request, "Please select at least one record.")
        return redirect("records:records_table")

    records = Record.objects.filter(RecordID__in=record_ids).order_by("-RecordID")

    deleted_count = 0
    request_count = 0
    skipped_deleted_count = 0
    skipped_pending_count = 0

    for rec in records:
        # CORRECTION: skip already deleted records
        if hasattr(rec, "IsDeleted") and rec.IsDeleted:
            skipped_deleted_count += 1
            continue

        # ============================================================
        # CORRECTION: Clerk path -> create one DELETE request per record
        # ============================================================
        if _is_clerk(user):
            existing_pending = _get_pending_delete_request_for_record(rec)
            if existing_pending:
                skipped_pending_count += 1
                continue

            req = ActionRequests.objects.create(
                RequestType="DELETE",
                TargetRecordID=rec,
                RequestedByUserID=user,
                RequestDetails=f"Bulk delete request for RecordID={rec.RecordID}",
                Status="pending",
                CreatedAt=timezone.now(),
            )

            _audit(
                request,
                "REQUEST_CREATE_DELETE_BULK_ITEM",
                f"Clerk requested bulk delete for RecordID={rec.RecordID} (RequestID={req.RequestID})",
                target_record=rec,
            )

            request_count += 1
            continue

        # ============================================================
        # CORRECTION: Admin path -> immediate soft delete
        # ============================================================
        _mark_record_deleted(rec, user=user)

        _audit(
            request,
            "RECORD_SOFT_DELETE_BULK_ITEM",
            f"Admin bulk soft-deleted RecordID={rec.RecordID}",
            target_record=rec,
        )

        deleted_count += 1

    # ============================================================
    # CORRECTION: final user feedback message
    # ============================================================
    if _is_clerk(user):
        if request_count > 0:
            messages.success(
                request,
                f"{request_count} delete request(s) submitted for admin approval."
            )
        elif skipped_pending_count > 0 or skipped_deleted_count > 0:
            messages.info(
                request,
                "No new delete requests were created. Selected records were already deleted or already had pending delete requests."
            )
        else:
            messages.error(request, "No delete requests were created.")

        return redirect("workflow:requests_list")

    if deleted_count > 0:
        msg = f"{deleted_count} record(s) moved to deleted records."
        if skipped_deleted_count > 0:
            msg += f" Skipped {skipped_deleted_count} already deleted record(s)."
        messages.success(request, msg)
    else:
        messages.info(request, "No records were deleted.")

    return redirect("records:records_table")


# ============================================================
# 8) attachment_upload_view (Clerk/Admin) — FIXED FOR YOUR SCHEMA
# ============================================================
@clerk_or_admin_required_view
def attachment_upload_view(request, record_id):
    user = request.current_user
    rec = get_object_or_404(Record, RecordID=record_id)

    if request.method != "POST":
        raise Http404()

    files = request.FILES.getlist("files")
    if not files:
        messages.error(request, "Please choose at least one file.")
        return redirect("records:record_detail", record_id=rec.RecordID)

    if not _ensure_media_root():
        messages.error(request, "MEDIA_ROOT is not configured. Cannot store attachments.")
        return redirect("records:record_detail", record_id=rec.RecordID)

    created = 0
    for f in files:
        rel_path = _save_uploaded_file(rec.RecordID, f)

        RecordAttachment.objects.create(
            RecordID_id=rec.RecordID,
            FilePath=rel_path,
            OriginalFileName=f.name,
            UploadedAt=timezone.now(),
            UploadedByUserID=user,
        )
        created += 1

    next_url = request.POST.get("next") or request.META.get("HTTP_REFERER")
    if next_url:
        return redirect(next_url)

    return redirect("records:record_detail", record_id=rec.RecordID)  # fallback


# ============================================================
# 9) attachment_download_view — FIXED FOR YOUR SCHEMA
# ============================================================
@login_required_view
def attachment_download_view(request, attachment_id):
    user = request.current_user
    att = get_object_or_404(RecordAttachment, AttachmentID=attachment_id)

    if hasattr(att, "IsDeleted") and att.IsDeleted and not _is_admin(user):
        return redirect("accounts:access_denied")

    # ✅ FIX: FK is RecordID_id
    rec = get_object_or_404(Record, RecordID=att.RecordID_id)
    if hasattr(rec, "IsDeleted") and rec.IsDeleted and not _is_admin(user):
        return redirect("accounts:access_denied")

    if not _ensure_media_root():
        raise Http404("MEDIA_ROOT is not configured.")

    rel_path = getattr(att, "FilePath", None)
    if not rel_path:
        raise Http404("No file path stored for this attachment.")

    abs_path = os.path.join(settings.MEDIA_ROOT, rel_path)
    if not os.path.exists(abs_path):
        raise Http404("File not found on disk.")

    filename = getattr(att, "OriginalFileName", None) or os.path.basename(rel_path)

    _audit(request, "ATTACHMENT_DOWNLOAD", f"Downloaded AttachmentID={att.AttachmentID} (RecordID={rec.RecordID})")
    return FileResponse(open(abs_path, "rb"), as_attachment=True, filename=filename)


def _image_to_pdf(image_path: Path, out_pdf_path: Path):
    c = canvas.Canvas(str(out_pdf_path), pagesize=A4)
    page_w, page_h = A4
    img = ImageReader(str(image_path))
    iw, ih = img.getSize()

    margin = 24
    max_w = page_w - margin * 2
    max_h = page_h - margin * 2
    scale = min(max_w / iw, max_h / ih)

    w = iw * scale
    h = ih * scale
    x = (page_w - w) / 2
    y = (page_h - h) / 2

    c.drawImage(img, x, y, width=w, height=h, preserveAspectRatio=True, mask="auto")
    c.showPage()
    c.save()


def _libreoffice_to_pdf(input_path: Path, out_dir: Path) -> Path:
    """
    Converts many formats to PDF using LibreOffice headless.

    IMPORTANT:
    - This prints the real error if conversion fails.
    - It tries both common Windows install paths.
    """
    out_dir.mkdir(parents=True, exist_ok=True)

    # ✅ Try common Windows install locations
    candidates = [
        r"C:\Program Files\LibreOffice\program\soffice.exe",
        r"C:\Program Files (x86)\LibreOffice\program\soffice.exe",
    ]

    soffice_path = None
    for c in candidates:
        if Path(c).exists():
            soffice_path = c
            break

    if not soffice_path:
        # This means LibreOffice isn't installed where we expect
        raise RuntimeError("LibreOffice soffice.exe not found in Program Files.")

    cmd = [
        soffice_path,
        "--headless",
        "--nologo",
        "--nolockcheck",
        "--nodefault",
        "--nofirststartwizard",
        "--convert-to",
        "pdf",
        "--outdir",
        str(out_dir),
        str(input_path),
    ]

    r = subprocess.run(cmd, stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True)

    # ✅ If it fails, raise the REAL error so you can see it
    if r.returncode != 0:
        raise RuntimeError(
            "LibreOffice failed.\n"
            f"CMD: {' '.join(cmd)}\n"
            f"STDOUT: {r.stdout}\n"
            f"STDERR: {r.stderr}\n"
        )

    produced = out_dir / (input_path.stem + ".pdf")
    if not produced.exists():
        raise RuntimeError("LibreOffice returned success but PDF was not produced.")
    return produced

@xframe_options_sameorigin
@login_required_view
def attachment_pdf_view(request, attachment_id):
    user = request.current_user
    att = get_object_or_404(RecordAttachment, AttachmentID=attachment_id)

    # Same visibility rules as download:
    if hasattr(att, "IsDeleted") and att.IsDeleted and not _is_admin(user):
        return redirect("accounts:access_denied")

    rec = get_object_or_404(Record, RecordID=att.RecordID_id)
    if hasattr(rec, "IsDeleted") and rec.IsDeleted and not _is_admin(user):
        return redirect("accounts:access_denied")

    if not _ensure_media_root():
        raise Http404("MEDIA_ROOT is not configured.")

    rel_path = getattr(att, "FilePath", None)
    if not rel_path:
        raise Http404("No file path stored for this attachment.")

    src_abs = Path(settings.MEDIA_ROOT) / rel_path
    if not src_abs.exists():
        raise Http404("File not found on disk.")

    ext = src_abs.suffix.lower()

    # Cache generated PDFs:
    cache_dir = Path(settings.MEDIA_ROOT) / "attachment_pdfs"
    cache_dir.mkdir(parents=True, exist_ok=True)
    cached_pdf = cache_dir / f"attachment_{att.AttachmentID}.pdf"

    # If cached already, serve it:
    if cached_pdf.exists():
        resp = FileResponse(open(cached_pdf, "rb"), content_type="application/pdf")
        resp["Content-Disposition"] = 'inline; filename="attachment.pdf"'
        return resp

    # If already a PDF, open original:
    if ext == ".pdf":
        resp = FileResponse(open(src_abs, "rb"), content_type="application/pdf")
        resp["Content-Disposition"] = 'inline; filename="attachment.pdf"'
        return resp

    # Images -> PDF:
    if ext in {".png", ".jpg", ".jpeg", ".webp"}:
        _image_to_pdf(src_abs, cached_pdf)
        resp = FileResponse(open(cached_pdf, "rb"), content_type="application/pdf")
        resp["Content-Disposition"] = 'inline; filename="attachment.pdf"'
        return resp

    # Other docs -> PDF using LibreOffice (required for “any file type”)
    try:
        produced = _libreoffice_to_pdf(src_abs, cache_dir)
        # Rename to stable cache name:
        if produced != cached_pdf:
            if cached_pdf.exists():
                cached_pdf.unlink()
            produced.rename(cached_pdf)

        resp = FileResponse(open(cached_pdf, "rb"), content_type="application/pdf")
        resp["Content-Disposition"] = 'inline; filename="attachment.pdf"'
        return resp
    except Exception:
        # If LibreOffice is not installed / conversion fails
        raise Http404("Cannot convert this file to PDF on this server. Install LibreOffice (soffice).")

# ============================================================
# 10) attachment_delete_view (Clerk/Admin) — FIXED FOR YOUR SCHEMA
# ============================================================
@clerk_or_admin_required_view
def attachment_delete_view(request, attachment_id):
    user = request.current_user
    
    if request.method != "POST":
        raise Http404()
    
    att = get_object_or_404(RecordAttachment, AttachmentID=attachment_id)
    rec_id = att.RecordID_id

    # Soft-delete if the model supports it, else hard-delete row + file
    if hasattr(att, "IsDeleted") or hasattr(att, "is_deleted") or hasattr(att, "deleted_at"):
        _mark_attachment_deleted(att, user=user)
        _audit(request, "ATTACHMENT_SOFT_DELETE", f"Soft-deleted AttachmentID={att.AttachmentID}")
    else:
        # attempt to delete file from disk
        if _ensure_media_root():
            rel_path = getattr(att, "FilePath", None)
            if rel_path:
                abs_path = os.path.join(settings.MEDIA_ROOT, rel_path)
                try:
                    if os.path.exists(abs_path):
                        os.remove(abs_path)
                except Exception:
                    pass

        att.delete()
        _audit(request, "ATTACHMENT_DELETE", f"Deleted AttachmentID={attachment_id}")

    messages.success(request, "Attachment removed.")
    
    next_url = request.POST.get("next") or request.META.get("HTTP_REFERER")
    if next_url:
        return redirect(next_url)
    
    return redirect("records:record_detail", record_id=rec_id)


# ============================================================
# 11) records_export_excel_view
# ============================================================
@login_required_view
def records_export_excel_view(request):
    user = request.current_user

    # CORRECTION: allow Export All when no filters are present.
    # If filters/sort exist, export filtered queryset.
    if _has_any_filter_params(request):
        qs = _build_filtered_records_qs(request, user)
        export_mode = "filtered"
    else:
        qs = Record.objects.all().order_by("-RecordID")

        # CORRECTION: keep "Export All" aligned with visible non-deleted records
        # unless include_deleted=1 is explicitly supplied by an admin.
        include_deleted = request.GET.get("include_deleted") == "1" and _is_admin(user)
        if not include_deleted:
            qs = qs.filter(**_record_is_deleted_filter())

        export_mode = "all"

    # limit export to a reasonable cap
    qs = qs[:5000]

    wb = Workbook()
    ws = wb.active
    ws.title = "MD_Registry_Export"

    headers = [
        "MessengerName",
        "Subject",
        "Description",
        "DateReceived",
        "InvoiceNumber",
        "IncomingDepartmentName",
        "OutgoingDepartmentName",
        "DateDispatched",
        "Status",
        "Returned",
        "DateReturned",
    ]
    ws.append(headers)

    def _fmt_dt(dt):
        if not dt:
            return ""
        try:
            return dt.astimezone(timezone.get_current_timezone()).strftime("%Y-%m-%d %H:%M")
        except Exception:
            try:
                return dt.strftime("%Y-%m-%d %H:%M")
            except Exception:
                return str(dt)

    def _fmt_date(d):
        if not d:
            return ""
        try:
            return d.strftime("%Y-%m-%d")
        except Exception:
            return str(d)

    def _dept_name(dept_fk):
        if not dept_fk:
            return ""
        return getattr(dept_fk, "DepartmentName", "") or ""

    for rec in qs:
        invoice_val = getattr(rec, "InvoiceNumber", "")
        try:
            invoice_str = f"{int(invoice_val):,}"
        except Exception:
            invoice_str = str(invoice_val) if invoice_val is not None else ""

        ws.append([
            getattr(rec, "MessengerName", "") or "",
            getattr(rec, "Subject", "") or "",
            getattr(rec, "Description", "") or "",
            _fmt_dt(getattr(rec, "DateReceived", None)),
            invoice_str,
            _dept_name(getattr(rec, "IncomingDepartmentID", None)),
            _dept_name(getattr(rec, "OutgoingDepartmentID", None)),
            _fmt_dt(getattr(rec, "DateDispatched", None)),
            getattr(rec, "Status", "") or "",
            getattr(rec, "Returned", "") or "",
            _fmt_date(getattr(rec, "DateReturned", None)),
        ])

    for idx in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(idx)].width = 24

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)

    filename = f"md_registry_export_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    # CORRECTION: audit message now reflects whether export was all or filtered
    if export_mode == "filtered":
        _audit(
            request,
            "RECORDS_EXPORT_EXCEL",
            f"Exported {qs.count()} record(s) to Excel (filtered export)."
        )
    else:
        _audit(
            request,
            "RECORDS_EXPORT_EXCEL",
            f"Exported {qs.count()} record(s) to Excel (all records export)."
        )

    resp = HttpResponse(
        bio.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp["Content-Disposition"] = f'attachment; filename="{filename}"'
    return resp

# ============================================================
# 12) records_import_excel_view — (kept as-is; safe via hasattr)
# ============================================================
@clerk_or_admin_required_view
def records_import_excel_view(request):
    user = request.current_user

    if request.method == "GET":
        return render(request, "records/records_import.html", {
            "is_admin": _is_admin(user),
            "is_clerk": _is_clerk(user),
        })

    file = request.FILES.get("file")
    if not file:
        messages.error(request, "Please upload an Excel file (.xlsx).")
        return redirect("records:records_import_excel")

    if not file.name.lower().endswith(".xlsx"):
        messages.error(request, "Invalid file type. Please upload a .xlsx file.")
        return redirect("records:records_import_excel")

    try:
        wb = load_workbook(file, data_only=True)
        ws = wb.active
    except Exception:
        messages.error(request, "Could not read the Excel file.")
        return redirect("records:records_import_excel")

    # -----------------------------
    # Header processing
    # -----------------------------
    headers = []
    for cell in ws[1]:
        headers.append((str(cell.value).strip() if cell.value is not None else ""))

    def _col(name: str):
        return headers.index(name) + 1 if name in headers else None

    # ✅ Required headers (schema aligned)
    required_headers = [
        "MessengerName",
        "Subject",
        "Description",
        "DateReceived",
        "InvoiceNumber",
        "IncomingDepartmentName",  # we recommend Name-based lookup
        "Status",
    ]

    missing = [h for h in required_headers if h not in headers]
    if missing:
        messages.error(request, f"Missing required column(s): {', '.join(missing)}")
        return redirect("records:records_import_excel")

    # Optional headers
    c_out_name = _col("OutgoingDepartmentName")
    c_date_disp = _col("DateDispatched")
    c_returned = _col("Returned")
    c_date_ret = _col("DateReturned")

    c_messenger = _col("MessengerName")
    c_subject = _col("Subject")
    c_desc = _col("Description")
    c_date_recv = _col("DateReceived")
    c_invoice = _col("InvoiceNumber")
    c_in_name = _col("IncomingDepartmentName")
    c_status = _col("Status")

    def _cell_str(row, col):
        if not col:
            return ""
        v = ws.cell(row=row, column=col).value
        return (str(v).strip() if v is not None else "")

    def _parse_excel_datetime(v):
        # Accept datetime object OR string (e.g. "2026-02-25 10:30" or ISO)
        if v is None or str(v).strip() == "":
            return None
        if isinstance(v, datetime):
            dt = v
        else:
            s = str(v).strip().replace("/", "-")
            try:
                dt = datetime.fromisoformat(s)
            except Exception:
                # try common "YYYY-MM-DD HH:MM"
                try:
                    dt = datetime.strptime(s, "%Y-%m-%d %H:%M")
                except Exception:
                    return None
        if timezone.is_naive(dt):
            try:
                dt = timezone.make_aware(dt, timezone.get_current_timezone())
            except Exception:
                pass
        return dt

    def _parse_excel_date(v):
        if v is None or str(v).strip() == "":
            return None
        if isinstance(v, date) and not isinstance(v, datetime):
            return v
        if isinstance(v, datetime):
            return v.date()
        s = str(v).strip().replace("/", "-")
        try:
            return date.fromisoformat(s)
        except Exception:
            return None

    def _normalize_status(s: str):
        s = (s or "").strip().lower()
        if s == "with md":
            return "With MD"
        if s == "not with md":
            return "Not with MD"
        return None

    def _normalize_yes_no(s: str):
        s = (s or "").strip().lower()
        if s == "yes":
            return "Yes"
        if s == "no":
            return "No"
        return ""

    created = 0
    failed = 0
    skipped_duplicates = 0
    errors = []

    # -----------------------------
    # Row-by-row import
    # -----------------------------
    for row_idx in range(2, ws.max_row + 1):
        # skip empty rows
        row_values = [ws.cell(row=row_idx, column=c).value for c in range(1, ws.max_column + 1)]
        if all(v is None or str(v).strip() == "" for v in row_values):
            continue

        messenger_name = _cell_str(row_idx, c_messenger)
        subject = _cell_str(row_idx, c_subject)
        description = _cell_str(row_idx, c_desc)

        # ✅ DateReceived required
        date_received_raw = ws.cell(row=row_idx, column=c_date_recv).value
        date_received = _parse_excel_datetime(date_received_raw)

        # ✅ InvoiceNumber required (commas allowed)
        invoice_raw = _cell_str(row_idx, c_invoice)
        invoice_clean = invoice_raw.replace(",", "").strip()
        invoice_int = None
        if invoice_clean:
            try:
                invoice_int = int(invoice_clean)
            except Exception:
                invoice_int = None

        # ✅ Incoming Department required (by name)
        incoming_name = _cell_str(row_idx, c_in_name)
        incoming_dept = _get_or_create_department_by_name(incoming_name)
        
        # ✅ Outgoing Department optional (auto-create if provided)
        outgoing_name = _cell_str(row_idx, c_out_name) if c_out_name else ""
        outgoing_dept = _get_or_create_department_by_name(outgoing_name) if outgoing_name else None

        # ✅ Status required and normalized
        status_raw = _cell_str(row_idx, c_status)
        status = _normalize_status(status_raw)

        # Optional fields
        outgoing_dept = _get_or_create_department_by_name(_cell_str(row_idx, c_out_name)) if c_out_name else None

        date_dispatched = None
        if c_date_disp:
            date_dispatched = _parse_excel_datetime(ws.cell(row=row_idx, column=c_date_disp).value)

        returned = _normalize_yes_no(_cell_str(row_idx, c_returned)) if c_returned else ""
        date_returned = _parse_excel_date(ws.cell(row=row_idx, column=c_date_ret).value) if c_date_ret else None

        # -----------------------------
        # ✅ Match Add/Edit validation rules
        # -----------------------------
        if not messenger_name:
            failed += 1
            errors.append(f"Row {row_idx}: MessengerName is required.")
            continue
        if not subject:
            failed += 1
            errors.append(f"Row {row_idx}: Subject is required.")
            continue
        if not description:
            failed += 1
            errors.append(f"Row {row_idx}: Description is required.")
            continue
        if not date_received:
            failed += 1
            errors.append(f"Row {row_idx}: DateReceived is required (use 'YYYY-MM-DD HH:MM').")
            continue
        if invoice_int is None:
            failed += 1
            errors.append(f"Row {row_idx}: InvoiceNumber must be numeric (commas allowed).")
            continue
        if not incoming_dept:
            failed += 1
            errors.append(f"Row {row_idx}: IncomingDepartmentName is required.")
            continue
        if not status:
            failed += 1
            errors.append(f"Row {row_idx}: Status must be 'With MD' or 'Not with MD'.")
            continue

        # If Date Dispatched => Outgoing Department required
        if date_dispatched and not outgoing_dept:
            failed += 1
            errors.append(f"Row {row_idx}: OutgoingDepartmentName is required when DateDispatched is filled.")
            continue

        # If dispatched AND With MD => Returned must be Yes and DateReturned required
        if date_dispatched and status == "With MD":
            if returned != "Yes":
                failed += 1
                errors.append(f"Row {row_idx}: Returned must be 'Yes' when Status is 'With MD' and DateDispatched is filled.")
                continue
            if not date_returned:
                failed += 1
                errors.append(f"Row {row_idx}: DateReturned is required when Status is 'With MD' and DateDispatched is filled.")
                continue

        # If Not with MD => clear returned fields (avoid stale data)
        if status == "Not with MD":
            returned = ""
            date_returned = None

        # -----------------------------
        # ✅ Duplicate detection (skip duplicates)
        # Key: InvoiceNumber + DateReceived(date) + IncomingDept + Subject
        # -----------------------------
        dup_exists = Record.objects.filter(
            InvoiceNumber=invoice_int,
            DateReceived__date=date_received.date(),
            IncomingDepartmentID=incoming_dept,
            Subject__iexact=subject,
        ).exists()

        if dup_exists:
            skipped_duplicates += 1
            continue

        # -----------------------------
        # Create record
        # -----------------------------
        try:
            rec = Record.objects.create(
                MessengerName=messenger_name,
                Subject=subject,
                Description=description,
                DateReceived=date_received,
                InvoiceNumber=invoice_int,
                IncomingDepartmentID=incoming_dept,
                OutgoingDepartmentID=outgoing_dept,
                DateDispatched=date_dispatched,
                Returned=returned if returned in ("Yes", "No") else None,
                DateReturned=date_returned,
                Status=status,
                CreatedAt=timezone.now() if hasattr(Record, "CreatedAt") else None,
                UpdatedAt=timezone.now() if hasattr(Record, "UpdatedAt") else None,
            )
            created += 1
        except Exception as e:
            failed += 1
            errors.append(f"Row {row_idx}: Could not create record ({str(e)[:140]}).")
            continue

    _audit(
        request,
        "RECORDS_IMPORT_EXCEL",
        f"Import complete. Created={created}, Failed={failed}, DuplicatesSkipped={skipped_duplicates}",
    )

    messages.success(
        request,
        f"Import complete: {created} created, {failed} failed, {skipped_duplicates} duplicates skipped."
    )

    if errors:
        messages.warning(request, "Some rows failed. First few errors: " + " | ".join(errors[:5]))

    return redirect("records:records_table")

# ============================================================
# 13) import_template_download_view
# ============================================================
@login_required_view
def import_template_download_view(request):
    user = request.current_user

    wb = Workbook()
    ws = wb.active
    ws.title = "ImportTemplate"

    # ✅ 11-field schema template headers (matches your Records model)
    headers = [
        "MessengerName",
        "Subject",
        "Description",
        "DateReceived",               # e.g. 2026-02-25 10:30
        "InvoiceNumber",              # commas allowed: 111,123
        "IncomingDepartmentName",     # REQUIRED (recommended)
        "OutgoingDepartmentName",     # optional unless DateDispatched filled
        "DateDispatched",             # optional datetime
        "Status",                     # REQUIRED: With MD / Not with MD
        "Returned",                   # conditional
        "DateReturned",               # conditional date: 2026-02-25
    ]
    ws.append(headers)

    # ✅ Example row
    ws.append([
        "John Messenger",
        "Budget Approval",
        "Document for MD review and approval.",
        "2026-02-25 10:30",
        "111,123",
        "Registry",
        "",                 # OutgoingDepartmentName
        "",                 # DateDispatched
        "With MD",
        "",                 # Returned
        "",                 # DateReturned
    ])

    for idx in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(idx)].width = 24

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)

    _audit(request, "IMPORT_TEMPLATE_DOWNLOAD", "Downloaded Excel import template (11-field schema)")

    filename = "md_registry_import_template.xlsx"
    resp = HttpResponse(
        bio.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp["Content-Disposition"] = f'attachment; filename="{filename}"'
    return resp